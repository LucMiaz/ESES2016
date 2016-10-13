import os
from pathlib import Path
import ast
from datetime import date, datetime
import openpyxl as xlsx
from openpyxl import load_workbook
import string
import py2neo as pyneo
from classes import *
import copy
import logging
import re

import urllib #for eniro api
def finddata30(OLDPATH):
    if OLDPATH.name=="Dropbox":
        newPath=OLDPATH.joinpath("ESES Course 2016/30 Data")
    elif OLDPATH.name=='30 Data':
        newPath=OLDPATH
    elif OLDPATH.name=='ESES Course 2016':
        newPath=OLDPATH.joinpath('30 Data')
    else:
        newPath=finddata30(OLDPATH.resolve().parents[0])
    return(newPath)
    
def data30(func, *args, **kwargs):
    def changePath(*args, **kwargs):
        OLDPATH=Path(os.getcwd())
        try:
            newPath=finddata30(OLDPATH)
        except IndexError:
            OLDPATH=Path("/Users/Shared/Dropbox/ESES Course 2016/")
            newPath=finddata30(OLDPATH)
        #print(newPath)
        os.chdir(newPath.as_posix())
        ret=func(*args,**kwargs)
        os.chdir(OLDPATH.as_posix())
        return(ret)
    return changePath
def sanitize_element(element):
    ele_=element
    ele_=ele_.replace(" ","_")
    ele_=ele_.replace("+","AND")
    ele_=ele_.replace("/","_")
    ele_=ele_.replace("*","STAR")
    ele_=ele_.replace("%","perc")
    ele_=ele_.replace("[","_")
    ele_=ele_.replace("]","_")
    ele_=ele_.replace("(","_")
    ele_=ele_.replace(")","_")
    ele_=ele_.replace("-","__")
    return(ele_)
def sanitize_headers(headers):
    newheaders=[]
    for ele in headers:
        if ele:
            newheaders.append(sanitize_element(ele))
        else:
            newheaders.append(ele)
    return(newheaders)
    
@data30
def import_xlsx(xlsxFileName, sheetNum=0, header=True, noBlankTitles=True, onlyHeaders=None,dataOnly=True, headersRow=0, sanitize=True):
    imported=[]
    wb2 = xlsx.load_workbook(filename=xlsxFileName, read_only=True, data_only=dataOnly)
    sheet=wb2[wb2.get_sheet_names()[sheetNum]]
    rows=sheet.rows
    for i in range(0,headersRow):
        next(rows)
    if header:
        headers=[cell.value for cell in next(rows)]
    else:
        headers=[]
        for i in range(0,len(rows[0])):
            headers.append('col'+i)
    copyheaders=[i for i in headers if i]
    copyheaders.sort()
    last=""
    count=1
    for i in copyheaders:
        if last==i:
            
            headers[headers.index(i)]=str(i)+str(count)
            count+=1
        else:
            count=1
        last=i
    if sanitize:
        headers=sanitize_headers(headers)
    for row in rows:
        rowD={}
        disposableHeaders=headers.copy()
        
        for cell in row:
            if len(disposableHeaders)>0:
                if disposableHeaders[0]:
                    rowD[disposableHeaders.pop(0)] = cell.value
                else:
                    disposableHeaders.pop(0)
        copyRow=copy.copy(rowD)
        
        imported.append(rowD)
    #sanitize de headers
    print(headers)
    return imported

def sheets_xlsx(xlsxFileName):
    wb2 = xlsx.load_workbook(filename=xlsxFileName, read_only=True)
    return wb2.get_sheet_names()

def write_xlsx(xlsxFileName, data):
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    for row in data:
        # Rows can also be appended
        ws.append(row)
    # Save the file
    wb.save(xlsxFileName)

def sanitize_sample_code(debut):
    original=debut
    if debut:
        if "+" in debut:
            divided=debut.split(".")#split the string around dots
            last=divided[-1]#takes the last element of the list
            start=divided[:-1]#take the first elements
            pluses=last.split("+")#split around +s
            plusret=[]
            for plus in pluses:
                if len(plus)==1:
                    plus="0"+plus#add a 0 at the start if the plus has only one digit
                plusret.append(plus)#adds it to the list of pluses
            last= "+".join(plusret)#join the plus again
            debut=".".join(start)+"."+last#joint the start and the end.
            print(original +" changed to "+ debut)
        elif len(debut.split(".")[-1])==1:
            last=debut[-1]
            debut=debut[:-1]+"0"+last
            print(original + " changed to "+ debut)
        elif not debut:
            debut=None
        else:##################################################################<---------------<-<-<-<-<-<
            pass
              
    return debut

def insert_PFxx():
    
    lc=match(label="Instrument", name="LC")
    if len(lc)>0:
        lc=lc[0]
    else:
        lc=Instrument(name="LC")
    folderName="39 PFxx"
    fileName="PFXX_161012_compilation_FaBa.xlsx"
    fill_PFxx(fileName=fileName, folderName=folderName, instrumentObj=lc) 
    

def fill_PFxx(fileName, instrumentObj, folderName="", idCode="sample ID", date="2016-09-23"):
    fileName=folderName+"/"+fileName
    idCode=sanitize_element(idCode)
    data=import_xlsx(xlsxFileName=fileName, headersRow=1, sanitize=True)
    date="2016-09-20"
    for row in data:
        if row[idCode]:
            ppcode=row['operator'].split(",")
            ppcodecopy=[]
            for pp in ppcode:
                ppcodecopy.append(pp.lower().replace(" ",""))
            ppcode=ppcodecopy
            del row['operator']
            obs=match(label="Observation", type="PFxx", ppcode=str(ppcode))
            if len(obs)>0:
                obs=obs[0]
            else:
                obs=Observation(type='PFxx', ppcode=str(ppcode),date=date)
                print("New observation ------------")
                for pp in ppcode:
                    person=match(label="Person", initials=pp)
                    if len(person)>0:
                        Relationship(obs, "BY", person[0])
                    else:
                        print(pp + " not found")
            units="ng/g"
            sampleId=row[idCode]
            try:
                print(sampleId)
            except:
                print("could not print id")
            sampleId=sanitize_sample_code(sampleId)
            print("Sanitized"+str(sampleId))
            patternBlank= re.compile('([B-b]lank)+(\s?)+([0-9])*\Z')
            patternControl= re.compile('([C-c]ontrol)+(\s?)+([0-9])*\Z')
            patternSample= re.compile('(E.[A-Z].)+([0-9]{2})+(.)+([0-9]{2})+([+]?[0-9]?[0-9]?)\Z')
            if sampleId:
                if re.match(patternBlank, sampleId):
                    sample=match(label="Blank", site="Blank", code=sampleId)
                    if sample:
                        if len(sample)>0:
                            sample=sample[0]
                    else:
                        sample=Sample(type="Blank", site="Blank", code=sampleId)
                        site=match(label="Site",code="Blank")[0]
                        Relationship(sample,"FROM", site)
                elif re.match(patternControl, sampleId):
                    sample=match(label="Control", site="Control", code=sampleId)
                    if sample:
                        if len(sample)>0:
                            sample=sample[0]
                    else:
                        sample=Sample(type="Control", site="Control", code=sampleId)
                        site=match(label="Site",code="Control")[0]
                        Relationship(sample,"FROM", site)
                elif re.match(patternSample,sampleId):
                    
                    sample=match(label='Sample', code=sampleId)
                    if sample:
                        if len(sample)>0:
                            sample=sample[0]
                        else:
                            print("did not find " + sampleId)
                            sample=Sample(type="ERROR", code=sampleId)
                    else:
                        print("problem with "+sampleId)
                        sample=Sample(type="ERROR", code=sampleId)
                    #siteCode=int(sampleId.split(".")[2])
                else:
                    sample=match(label="Sample", type="Sediment", code=row[idCode])
                    if len(sample)>0:
                        sample=sample[0]
                    else:
                        sample=Sample(type='Sediment',site='Archive',code=row[idCode], siteCode="Archive")
                        site=match(label="Site",code="Archive")[0]
                        Relationship(sample,"FROM", site)
                        print("Created rel for "+row[idCode])

            del row[idCode]
            row['type']="PFxx"
            row['units']=units
            rel=match(label="OF", relationshipNode=True, target=sample, origin=obs)
            if rel:
                if len(rel)>0:
                    print('rel found')
                    rel=rel[0]
                    rel.add(**row)
                    rel.push()
                else:
                    print('no match for relationship')
                    Relationship(obs,"OF",sample,**row)
            else:
                print('match :none')
                Relationship(obs,"OF",sample,**row)


def insert_data_from_file(fileName, relationshipProps, instrumentToValue, observationObj, headersRow=0,folderName="", idLabel="Sample Id", date=None, sheetNum=0):
    """
    indicate how to find the sample id
    indicate how to find the ppcode
    relationshipProps must contain a list of duples with the original label, the destinated label, unit
    
    """

    fileName=folderName+fileName
    logger.info('reading : %s', fileName)
    data=import_xlsx(xlsxFileName=fileName, headersRow=headersRow, sanitize=True, sheetNum=sheetNum)
    relProps=[]
    idLabel=sanitize_element(idLabel)
    for original, target, unit in relationshipProps:
        relProps.append([sanitize_element(original), sanitize_element(target), unit])
    dataObjects=[]
    logger.debug('Will look for: %s', relProps)
    for datum in data:
        props={}
        sprops={}
        units={}
        sampleId=datum[idLabel]
        for original, target,unit in relProps:
            try:
                props[target]=datum[original]
            except:
                logger.info("No "+original+" in "+ str(sampleId))
            if unit:
                units[target]=unit
        props["units"]=units
        logger.info('Matching sample ...')
        sample=match(label="Sample", code=sampleId)
        if len(sample)>0:
            logger.info('Sample matched: %'+str(sample[0]))
            sample[0].add(**props)
            r=Relationship(observationObj, "OF",sample[0], **props)
            dataObjects.append(r)
            sample=sample[0]
        else:
            if sampleId=="SE3":
                sampleType="Sediment"
            elif sampleId=="tom" or sampleId=="Tom" or sampleId=="Blank" or sampleId=="blank":
                sampleType="Blank"
            elif sampleId== None:
                sampleType=None
            elif len(sampleId.split('.'))>1:
                if sampleId.split('.')[1]=="S":
                    sampleType="Sediment"
                elif sampleId.split('.')[1]=="W":
                    sampleType="Water"
                elif sampleId.split('.')[1]=="Z":
                    sampleType="Zoo"
                elif sampleId.split('.')[1]=="F":
                    sampleType="Fish"
                else:
                    sampleType="Unknown"
            else:
                sampleType="Sediment"
            if sampleType:
                sample=Sample(code=sampleId, type=sampleType)
                r=Relationship(observationObj, "OF",sample, **props)
                dataObjects.append(r)
                logger.info('No samples found.')
    
    
    # update records here
    logger.info('Finished reading file')
    return dataObjects

def insert_Hg():
    folderName="zzz_37 HgDMA_treated"
    fileNames=["HgDMA_160920_ESES1_sed(1)OD+zoo prel MaMe.xlsx","HgDMA_160922_ESES2_sed(1)OD+zoo prel MaMe.xlsx","HgDMA_160928_ESES_fish+sed(1)ODFD prel2 MaMe.xlsx","HgDMA_160929_ESES_sed(1a)FD prel MaMe.xlsx","HgDMA_160930_ESES_sed(1b)FD prel MaMe.xlsx","HgDMA_161007_ESES3_sed(2)FD prel MaMe.xlsx"]
    #fileNames=["HgDMA_161007_ESES_sed(2)FD prel MaMe.xlsx"]
    dma=match(label="Instrument",name="DMA-80")
    if len(dma)>0:
        dma=dma[0]
    else:
        dma=Instrument(name="DMA-80",manufacturer="Milestones Srl", website="http://www.milestonesrl.com/en/mercury/dma-80/features.html")
    groups={"ESES1":["lumi","faba","jeis","cagr","lojs","giho"],"ESES2":["erwi","mamä","idbo","mahe","laan","jojo"],"ESES3":["mahe"], "ESES4":["lumi","faba"], "ESES5":["lumi","faba","mahe"], "ESES6":[]}
    for filen in fileNames:
        print(filen)
        xlsxFileName= folderName+"/"+filen
        relationshipProps=[["W boat","boatWg","g"],["W sample","sampleWg","g"],["W boat+ash","boatANDashWg","g"],["Comments","comments",None],["Hg ng","Hg","ng"],["Hg ng/g","Hg_","ng/g"],["Row","row",None],['drying','drying',None],['error','error',None]]
        splitted=filen.split("_")
        date=splitted[1]
        group=splitted[2]
        name=splitted[3]
        d=import_xlsx(xlsxFileName)
        try:
            ppcodes=d[0]['ppcode']
        except:
            if group in groups.keys():
                ppcodes=groups[group]
            else:
                ppcodes=None
                print("No ppcodes for "+xlsxFileName)    
            
        
        observationObj=match(label="Observation", name=name, date=date)
        if len(observationObj)>0:
            observationObj=observationObj[0]
        else:
            observationObj=Observation(type="Hg", name=name, date=date, group=group)
            Relationship(observationObj,"WITH",dma)
            if ppcodes:
                for person in ppcodes:
                    personObj=match(label="Person", initials=person.lower().replace(" ",""))
                    if len(personObj)>0:
                        personObj=personObj[0]
                        Relationship(observationObj, "BY", personObj, group=group, date=date)
        dataObj=insert_data_from_file(xlsxFileName, relationshipProps,dma,observationObj, headersRow=0, idLabel="ID sample")
        if len(dataObj)>0:
            for obj in dataObj:
                if "boatWg" in obj.__dict__.keys() and "sampleWg" in obj.__dict__.keys() and "boatANDashWg" in obj.__dict__.keys():
                #if sampleId in obj.__dict__.keys():
                    if obj.sampleWg>0:
                        loi=1-(obj.boatANDashWg-obj.boatWg)/obj.sampleWg
                        print(str(loi))
                        obj.add(loi=loi)
                        units=obj.units
                        units['loi']=loi
                        obj.add(units=units)
                        obj.push()
            
#TODO
"""
def insert_water():
    filename="33 Water/161010_Water compilation working sheet_CaGr.xlsx"
    turbidity=import_xlsx(filename, 
"""
    
@data30
def insert_sites(xlsxFileName="31 Sites/Site description (Sunbeam).xlsx"):
    data=import_xlsx(xlsxFileName=xlsxFileName, sheetNum=2, sanitize=False, headersRow=1)
    letters=list(string.ascii_uppercase)
    sitesObj=[]
    for site in data:
        
        if site['GPS'] and isinstance(site['Site'],int):
            latraw,lonraw=site["GPS"].split(",")
            lat=float(latraw[2:4])+float(latraw[6:12])/60
            lon=float(lonraw[3:5])+float(lonraw[7:13])/60
            
            siteDict={"name":site["Site name"], "code":site["Site"], "letter":letters.pop(0), "lat":round(lat,6), "lon":round(lon,6),"temperature":site["Temperature (°C)"], "depth":site["Depth (m)"],"region":site["Region"], "eniroDepth":site["Eniro depth (m)"], "regionSed":site['regionSediments'],"regionWater":site['regionWater']}
            matched=match(label="Site", code=site["Site"], letter=siteDict['letter'], lat=siteDict['lat'], lon=siteDict['lon'])
            if len(matched)>0:
                sitesObj.append(matched[0])
            else:
                sitesObj.append(Site(**siteDict))
            
    return sitesObj
    
@data30   
def insert_ctd(xlsxName='33 Water/332 CTD/CTD_160926_rawData_MaHe.xlsx'):
    #inserting ctd observations
    ctdHeaders=sheets_xlsx(xlsxName)
    ctdInst=Instrument(name="ctd", measures=[{"pressure":"dbar"},{"temperature":"degC"},{"conductivity":"mS/cm"},{"salinity":"ppt"},{"sigma":"kg/m3"},{"time":"absolute"}])
    for sheetNum in ctdHeaders:
        try:
            siteCode=(sheetNum.split("_")[0]).split(" ")[1]
        except:
            if siteCode == "Transect":
                siteCode=0
                site=Site(code=0, name="Transect", letter="T")
        else:
            try:
                siteCode="".join([letter if letter in ["0","1","2","3","4","5","6","7","8","9"] else "" for letter in siteCode])
            except:
                print("Problem with "+str(siteCode)+". Could not digit-join.")
            else:
                if siteCode=="":
                    siteCode=99
        try:
            siteCode=int(siteCode)
        except:
            print("Could not convert "+str(siteCode)+ " to string")
        print(siteCode)
        try:
            repCode=sheetNum.split("_")[1]
        except:
            failed.append(str(sheetNum))
        else:
            try:
                repCode=int(repCode)
            except ValueError:
                repCode="".join([letter if letter in ["0","1","2","3","4","5","6","7","8","9"] else "" for letter in repCode])
                if repCode=="":
                    repCode=99
                repCode=int(repCode)
            code="E.ctd."+str(siteCode)+"."+str(repCode)
            imported=import_xlsx(xlsxName,sheetNum=ctdHeaders.index(sheetNum), sanitize=False)
            dateObs= imported[0]["date"]
            values=[sheetNum,repCode,str(dateObs.year)+"-"+str(dateObs.month)+"-"+str(dateObs.day),siteCode]
            failed=[]
            #
            ctd=Sample(type="ctd",site=siteCode, number=repCode, date=str(dateObs.year)+"-"+str(dateObs.month)+"-"+str(dateObs.day))
            instrumentrelation=Relationship(ctd,"TAKEN_WITH",ctdInst)
            
            data={"dataset":[],"pressure":[],"temperature":[],"conductivity":[],"salinity":[],"sigma":[],"time":[]}
            dataUnits={"dataset":"","pressure":"dbar","temperature":"degC","conductivity":"mS/cm","salinity":"ppt","sigma":"kg/m3","time":"absolute"}
            for datum in imported:
                time=datum["time"]
                timeStr=str(time.hour)+":"+str(time.minute)
                measure={"dataset":datum["Datasets"],"pressure":datum["Press [dbar]"],"temperature":datum["Temp [degC]"],"conductivity":datum["Cond [mS/cm]"],"salinity":datum["SALIN [ppt]"],"sigma": datum["SIGMA [kg/m3]"],"time":timeStr}
                data["dataset"].append(datum["Datasets"])
                data["pressure"].append(datum["Press [dbar]"])
                data["temperature"].append(datum["Temp [degC]"])
                data["conductivity"].append(datum["Cond [mS/cm]"])
                data["salinity"].append(datum["SALIN [ppt]"])
                data["sigma"].append(datum["SIGMA [kg/m3]"])
                data["time"].append(timeStr)
                #testing if up and down
                
                #obs=Observation(**measure)
                #Relationship(obs,"OF",ctd)
                #Relationship(obs,"BY",ctdInst)
            lowest=max(data['pressure'])
            highest=min(data['pressure'])
            lindex=data['pressure'].index(lowest)
            half=data['pressure'][lindex:]
            halfhighest=min(half)
            if (halfhighest-lowest)/(highest-lowest) < 0.5:
                data["run"]="down"
            else:
                data["run"]="down+up"
            data['code']=code
            ctd.add(**data)
            ctd.push()
            site=match(label="Site",code=int(siteCode))
            if len(site)>0:
                Relationship(ctd,"TAKEN_IN",site[0])
            else:
                print("No site to match for ctd measure : "+str(ctd))
    return failed




def insert_people():
    People=[]
    people=[["Luc","Miaz","luc.miaz@gmail.com"],["Lotta", "Sjöholm", "lotta.sjoholm@live.com"],["Erika", "Wikmark", "erikawikmark@gmail.com"],["Magdalena", "Mähler", "magdalenamahler@hotmail.com"],["Ida", "Bohman", "ida.c.bohman@gmail.com"],["Fabian", "Balk", "fabian.g.p.balk@gmail.com"],["Gisela", "Horlitz", "gisela_horlitz@gmx.de"],["Markus", "Hermann", "markushermann01@googlemail.com"], ["Jennifer", "Isaksson", "jennifer.isaksson@gmail.com"],["Caroline", "Grannemann", "caroline.grannemann@rwth-aachen.de"], ["Laura", "Anthony", "laura_anthony@hotmail.com"],["Josefine", "Johansson", "josefine.johansson@hotmail.com"]]
    for person in people:
        newone=match(label="Person",name=person[0],surname=person[1])
        if len(newone)>0:
            People.append(newone[0])
        else:
            newone=Person(name=person[0],surname=person[1],email=person[2])
            People.append(newone)
    return(People)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@data30
def insert_sediments(xlsxFileName="34 Sediment/Sediment_161010_compilation_lumi.xlsx",auto=False, rounding=3):
    sheetNames=sheets_xlsx(xlsxFileName)
    if auto:
        sheetNum=1
    else:
        print("The file "+str(xlsxFileName)+" contains the following sheets :")
        for i in range(0,len(sheetNames)):
            print(str(i+1)+". "+sheetNames[i])
        res=0
        while res < 1 or res-1>len(sheetNames):
            res=input("Select sheet name (use q to quit): ")
            if res=="q":
                res=1
            else:
                try:
                    res=int(res)
                except:
                    print("insert only integers or q")
                    res=0
        sheetNum=res-1
    data=import_xlsx(xlsxFileName=xlsxFileName,sheetNum=sheetNum,dataOnly=True, sanitize=False)
    priorityraw=import_xlsx(xlsxFileName=xlsxFileName, sheetNum=sheetNum+1, dataOnly=True, sanitize=False)
    for row in priorityraw:
        codepriority=row['Sample code']
        print(codepriority)
        try:
            res=codepriority.split(".")
        except:
            pass
        else:
            if len(res)>1:
                if res[0]=="E" and res[1]=="S":
                    sample=match(label="Sample", type="Sediment", code=codepriority)
                    if len(sample)>0:
                        props={"priority":1,"color":row["Colour"], "grainSize":row["Grain size"],"OtherComments":row["Other comments"]}
                        sample[0].add(**props)
                        sample[0].push()
    #todo add type of water in the sediment xlsx and import it
    sediments=[]
    for sample in data:
        
        labels={}
        variables={}
        units={}
        labelsToVar=[
                    ['Cermic vial weight + sample','ceramicAndSampleWg','g'],
                    ['After burning','ceramicAfterBurnWg','g'],
                    ['Weight for PF (g)','PFusedWg','g'],
                    ['weight to get 1mg of organic matter [mh]','onemgOrgMatter','mg'],
                    ['Ceramic vial weight','ceramicWg','g'],
                    ['Wet-weight sample (g)','wetSampleWg','g'],
                    ['Weight Aluminium + sample (g)','aluminiumAndSampleWg','g'],
                    ['Weight Aluminium box (g)','aluminiumWg','g'],
                    ['Dry weight sample (g)','drySampleWg','g'],
                    ['DC [%]','dc','%'],
                    ['Ceramic vial code','ceramicCode',''],
                    ['Sample code','code','g'],
                    ['Dry weight sample aluminium box+sample (g)','drySampleAndAluminiumWg','g'],
                    ['Weight of plastic bottle (g)','plasticBottleWg','g']]
        unit='gram'
        unitabb='g'
        
        #TODO waterType
        seds=match(label="Sample", type="Sediment", code=sample['Sample code'])
        if len(seds)>0:
            seds=seds[0]
            print("found "+ sample['Sample code'])
        else:
            seds=Sample(type="Sediment", code=sample['Sample code'])
            print("Not found : "+sample['Sample code'])
        #seds.add(priority=prio)
        for ltv in labelsToVar:
            labels[ltv[0]]=ltv[1]
            variables[ltv[1]]=ltv[0]
            units[ltv[1]]=ltv[2]
            valueofprop=sample[ltv[0]]
            if isinstance(valueofprop,float):
                valueofprop=round(valueofprop,rounding)
            try:
                propsToAdd={ltv[1]:valueofprop}
            except:
                print(ltv[0])
            seds.add(**propsToAdd)
        loi=1-(seds.drySampleAndAluminiumWg-seds.aluminiumWg)/(seds.aluminiumAndSampleWg-seds.aluminiumWg)
        wc=(seds.wetSampleWg-seds.drySampleWg)/seds.wetSampleWg
        units["loi"]="ratio of dw"
        units['wc']='ratio of ww'
        props2add={"units":units, "loi":loi, "wc":wc}
        seds.add(**props2add)
        #print(str(loi))
        seds.push()
        sediments.append(seds)
    
    return(sediments)



    

